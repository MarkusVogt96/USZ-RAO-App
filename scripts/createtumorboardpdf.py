import UNIVERSAL
import os
import sys
import time
import pyautogui
import clipboard
from datetime import datetime
import shutil
import re
import json
import openpyxl
from openpyxl import load_workbook
import fitz  # PyMuPDF
import io
from PIL import Image
import UNIVERSAL # Ihre eigene Modul-Datei


# Get the absolute path of the directory containing the current script (patdata.py)
script_dir = os.path.dirname(os.path.abspath(__file__))
# Get the absolute path of the parent directory (the application root, e.g., "USZ RAO App")
app_dir = os.path.dirname(script_dir)
# Define the path to the main screenshots directory (located within the script directory)
screenshots_dir = os.path.join(script_dir, 'screenshots pyautogui')
local_screenshots_dir = os.path.join(screenshots_dir, "tumorboards")
# Define the absolute path to patdata in the user's home directory
user_home = os.path.expanduser("~")


#Definne global variables
alle_tumorboards = [
    "GI",
    "Gyn Becken",
    "Gyn Mamma",
    "HCC",
    "HPB",
    "Melanom",
    "Neuro",
    "NET",
    "KHT",
    "Pädiatrie",
    "Hypophyse",
    "Lymphom",
    "Sarkom",
    "Schädelbasis",
    "Schilddrüse",
    "Thorax",
    "Uro"
]

tumorboard_auswahl = None


tumorboard_png_mapping = {
    "GI": "button_tumorboard_gi.png",
    "Gyn Becken": "button_tumorboard_gynbecken.png",
    "Gyn Mamma": "button_tumorboard_XXX.png",
    "HCC": "button_tumorboard_hcc.png",
    "HPB": "button_tumorboard_XXX.png",
    "Melanom": "button_tumorboard_melanom.png",
    "Neuro": "button_tumorboard_neuro.png",
    "NET": "button_tumorboard_net.png",
    "KHT": "button_tumorboard_kht.png",
    "Pädiatrie": "button_tumorboard_XXX.png",
    "Hypophyse": "button_tumorboard_XXX.png",
    "Lymphom": "button_tumorboard_llmz.png",
    "Sarkom": "button_tumorboard_sarkom.png",
    "Schädelbasis": "button_tumorboard_schaedelbasis.png",
    "Schilddrüse": "button_tumorboard_XXX.png",
    "Thorax": "button_tumorboard_thorax.png",
    "Uro": "button_tumorboard_uro.png"
}

def userinput():
    print("Bitte wählen Sie ein Tumorboard aus:")
    for idx, tb in enumerate(alle_tumorboards, 1):
        print(f"{idx}: {tb}")
    while True:
        try:
            auswahl = int(input("Nummer eingeben: "))
            if 1 <= auswahl <= len(alle_tumorboards):
                global tumorboard_auswahl
                tumorboard_auswahl = alle_tumorboards[auswahl - 1]
                print(f"Ausgewählt: {tumorboard_auswahl}")
                # Set the path to the tumorboard folder
                global tb_folder
                # Check if the network path exists
                net_path = r"K:\RAO_Projekte\App"
                if os.path.exists(net_path):
                    tb_folder = os.path.join(net_path, "tumorboards", tumorboard_auswahl)
                else:
                    tb_folder = os.path.join(os.path.expanduser("~"), "tumorboards", tumorboard_auswahl)
                break
            else:
                print("Ungültige Nummer. Bitte erneut versuchen.")
        except ValueError:
            print("Bitte eine gültige Zahl eingeben.")

def open_tumorboard():
    pyautogui.press('f6') #öffnet Briefkasten
    if not UNIVERSAL.find_button("button_aufgaben_confirm.png", base_path=local_screenshots_dir, confidence=0.9, max_attempts=100): 
        print("Fehler: Konnte button_aufgaben_confirm.png nicht finden. Briefkasten wurde nicht geöffnet. Beende Skript.")
        sys.exit(1)
    time.sleep(0.5)

    pyautogui.click(740, 120) #klickt auf Briefkasten

    while True:
        if UNIVERSAL.find_button(image_name="button_briefkasten_confirm.png", base_path=local_screenshots_dir, max_attempts=1, confidence=0.9): 
            print("button_briefkasten_confirm.png gefunden, break While loop.")
            break
        elif UNIVERSAL.find_and_click_button(image_name="button_briefkasten.png", base_path=local_screenshots_dir, max_attempts=1, confidence=0.9): 
            print("button_briefkasten.png gefunden und geklickt, break While loop.")
            break
        else:
            print("Fehler: button_briefkasten.png und button_briefkasten_confirm.png nicht gefunden, loope weiter.")
            time.sleep(0.1)
            continue

    if not UNIVERSAL.find_and_click_button(image_name="button_suchleiste.png", base_path=local_screenshots_dir): print("Fehler: Konnte x-Suchleiste-Button nicht klicken."); sys.exit()
    tumor = "tumor"
    clipboard.copy(tumor)
    pyautogui.hotkey('ctrl', 'v') #fügt den Text "tumor" ein

    tumorboard_png = tumorboard_png_mapping.get(tumorboard_auswahl)
    print(f"tumorboard_png: {tumorboard_png}")
    if tumorboard_png is None:
        print(f"Fehler: Tumorboard {tumorboard_auswahl} nicht gefunden.")
        sys.exit(1)
    if not UNIVERSAL.find_and_click_button(image_name=tumorboard_png, base_path=local_screenshots_dir, confidence=0.95): print("Fehler: Konnte spez. Tumorboard-Button nicht klicken."); sys.exit()

    time.sleep(2)

    if UNIVERSAL.find_button("button_keine_patienten.png", base_path=local_screenshots_dir, confidence=0.95, interval=0.1, max_attempts=10): 
        print("ACHTUNG: Zum aktuellen Zeitpunkt keine Patienten für das ausgewählte Tumorboard angemeldet! Breche ab.")
        # Erstelle eine .txt-Datei im tumorboard-Ordner mit aktuellem Datum und Hinweis
        
        os.makedirs(tb_folder, exist_ok=True)
        date_str = datetime.now().strftime("%d.%m.%Y")
        file_path = os.path.join(tb_folder, f"{date_str}_KEINE_PATIENTEN.txt")
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(f"Aktuell sind keine Patienten für das Tumorboard {tumorboard_auswahl} angemeldet.\n")
        return

    if not UNIVERSAL.find_and_click_button_offset(image_name="button_erledigt_am.png", base_path=local_screenshots_dir, y_offset=30): print("Fehler: freier_bereich_Button nicht gefunden."); return False
    if not UNIVERSAL.find_button("button_zeile_unter_erledigt_markiert.png", base_path=local_screenshots_dir, confidence=0.95, max_attempts=80): 
        print("Fehler: button_zeile_unter_erledigt_markiert.png nicht gefunden. Beende Skript."); return False
    pyautogui.hotkey('ctrl', 'a') #wählt alles aus
    if not UNIVERSAL.find_button("button_alle_markiert_confirm.png", base_path=local_screenshots_dir, confidence=0.95, max_attempts=80): 
        print("Fehler: button_alle_markiert_confirm.png nicht gefunden. Beende Skript."); return False

def tumorboard_nach_nachname_sortieren():
    if not UNIVERSAL.find_and_click_button_offset(image_name="button_erledigt_am.png", base_path=local_screenshots_dir, y_offset=30, rightclick=True):
        print("Konnte nicht unter Erledigt am rechtsklicken.")
        return False
    
    if not UNIVERSAL.find_and_click_button("button_sortieren.png", base_path=local_screenshots_dir, confidence=0.95, max_attempts=80): print("Fehler: Konnte button_sortieren.png nicht klicken."); return False

    UNIVERSAL.ctrl_tabs(2) #Wechselt in den Tab "Spalte 1"
    pyautogui.press('enter') #Bestätigen
    pyautogui.typewrite("pat") #pat schreiben
    time.sleep(0.3)
    pyautogui.press('enter') #wählt Patienten aus
    time.sleep(0.3)

    UNIVERSAL.ctrl_tabs(1)
    pyautogui.press('enter') #wählt Art der Sortierung aus
    pyautogui.typewrite("auf") #auf schreiben
    pyautogui.press('enter') #wählt aufsteigend aus
    time.sleep(0.3)
    if not UNIVERSAL.find_and_click_button("button_sortierung_ok.png", base_path=local_screenshots_dir, confidence=0.95): print("Fehler: Konnte Button_sortierung_ok.png nicht klicken."); return False


def als_pdf_speichern():
    if not UNIVERSAL.find_and_click_button_offset(image_name="button_erledigt_am.png", base_path=local_screenshots_dir, y_offset=30, rightclick=True):
        print("Konnte nicht unter Erledigt am rechtsklicken.")
        return False
    
    if not UNIVERSAL.find_and_click_button("button_aufgaben_drucken.png", base_path=local_screenshots_dir, confidence=0.95, max_attempts=80): print("Fehler: Konnte button_als_pdf_speichern.png nicht klicken."); return False

    if not UNIVERSAL.find_and_click_button_offset("button_drucker.png", base_path=local_screenshots_dir, confidence=0.95, x_offset=170): print("Fehler: Konnte button_drucker.png nicht klicken."); return False
    if not UNIVERSAL.find_button("button_druckeinrichtung_confirm.png", base_path=local_screenshots_dir, confidence=0.95, max_attempts=80): 
        print("Fehler: button_druckeinrichtung_confirm.png nicht gefunden. Beende Skript."); return False

    if not UNIVERSAL.find_and_click_button_offset("button_name.png", base_path=local_screenshots_dir, confidence=0.95, x_offset=150): print("Fehler: Konnte button_name.png nicht klicken."); return False
    if not UNIVERSAL.find_and_click_button("button_microsoft_print_to_pdf.png", base_path=local_screenshots_dir, confidence=0.95): print("Fehler: Konnte button_microsoft_print_to_pdf.png nicht klicken."); return False
    if not UNIVERSAL.find_button("button_print_to_pdf_confirm.png", base_path=local_screenshots_dir, confidence=0.95, max_attempts=80): 
        print("Fehler: button_print_to_pdf_confirm.png nicht gefunden. Beende Skript."); return False
    
    time.sleep(0.4)
    pyautogui.press('enter') #Bestätigen
    time.sleep(2)
    pyautogui.press('enter') #Bestätigen
    
    # Try-Loop: Bis zu 80 Versuche, jeweils beide Optionen mit max_attempts=1 prüfen
    for attempt in range(80):
        if UNIVERSAL.find_and_click_button("button_ja_rot.png", base_path=local_screenshots_dir, confidence=0.90, max_attempts=1): 
            print("button_ja_rot.png wurde gefunden. Break loop.")
            break
        if UNIVERSAL.find_button("button_druckausgabe_speichern_unter.png", base_path=local_screenshots_dir, confidence=0.90, max_attempts=1):
            print("button_druckausgabe_speichern_unter.png wurde gefunden. Break loop.")
            break
        time.sleep(0.05)
    else:
        print("Fehler: Weder button_ja_rot.png noch button_druckausgabe_speichern_unter.png gefunden nach 80 Versuchen.")
        return False
    
    # PDF-Speichern-Loop für alle Tumorboard-Anmeldungen
    #Export dir definieren
    if not UNIVERSAL.find_and_click_button("button_export_zeile.png", base_path=local_screenshots_dir, confidence=0.90): print("Fehler: Konnte button_export_zeile.png nicht klicken."); sys.exit()
    clipboard.copy(tb_folder)
    pyautogui.hotkey('ctrl', 'v') #fügt den Text "Export" ein
    pyautogui.press('enter') #Bestätigen
    time.sleep(0.3)

    nummer = 1
    while UNIVERSAL.find_and_click_button(
        "button_dateiname.png",
        base_path=local_screenshots_dir,
        confidence=0.90,
        max_attempts=50,
        interval=0.1
    ):
        pyautogui.typewrite(str(nummer))
        time.sleep(0.1)
        pyautogui.press("enter")
        nummer += 1
    print("Alle Tumorboard-Anmeldung erfolgreich als PDFs einzeln exportiert")

    # Erstelle einen neuen Ordner mit aktuellem Datum im tb_folder
    date_str = datetime.now().strftime("%d.%m.%Y")
    global dated_folder
    dated_folder = os.path.join(tb_folder, date_str)
    if os.path.exists(dated_folder):
        # Lösche alle Inhalte im Ordner
        for filename in os.listdir(dated_folder):
            file_path = os.path.join(dated_folder, filename)
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
    else:
        os.makedirs(dated_folder, exist_ok=True)

    # Verschiebe alle PDFs mit nur Zahlen im Namen in den neuen Ordner
    for filename in os.listdir(tb_folder):
        if filename.lower().endswith(".pdf") and re.fullmatch(r"\d+\.pdf", filename):
            src = os.path.join(tb_folder, filename)
            dst = os.path.join(dated_folder, filename)
            shutil.move(src, dst)
    print(f"Alle exportierten PDFs wurden nach {dated_folder} verschoben.")



def pdfs_umbenennen(dated_folder):
    import fitz  # PyMuPDF

    # Regex für 5-10-stellige Zahl
    patient_num_re = re.compile(r"\b\d{5,10}\b")

    for filename in os.listdir(dated_folder):
        if filename.lower().endswith(".pdf") and re.fullmatch(r"\d+\.pdf", filename):
            pdf_path = os.path.join(dated_folder, filename)
            try:
                with fitz.open(pdf_path) as doc:
                    text = ""
                    for page in doc:
                        text += page.get_text()
                # Suche nach "Patienteninformationen" und dann nach erster 5-10-stelliger Zahl danach
                idx = text.find("Patienteninformationen")
                patientennummer = None
                if idx != -1:
                    # Nur Text nach "Patienteninformationen" durchsuchen
                    after = text[idx:]
                    match = patient_num_re.search(after)
                    if match:
                        patientennummer = match.group()
                if patientennummer:
                    new_filename = f"{os.path.splitext(filename)[0]} - {patientennummer}.pdf"
                    new_path = os.path.join(dated_folder, new_filename)
                    os.rename(pdf_path, new_path)
                    print(f"{filename} umbenannt in {new_filename}")
                else:
                    print(f"Keine Patientennummer in {filename} gefunden.")
            except Exception as e:
                print(f"Fehler beim Auslesen von {filename}: {e}")


def create_excel_file():
    import fitz  # PyMuPDF
    import io
    from PIL import Image
    import UNIVERSAL # Eigene Modul-Datei

    # Regex patterns
    patient_num_re = re.compile(r"\b\d{5,10}\b")
    geschlecht_geb_re = re.compile(r"\b([MW])\s*/\s*(\d{2}\.\d{2}\.\d{4})")

    # Collect data for each PDF
    patient_data = []

    # Sortiere Dateien numerisch. Funktioniert für "1 - 12345.pdf"
    pdf_files = sorted(
        [f for f in os.listdir(dated_folder) if f.lower().endswith(".pdf") and re.fullmatch(r"\d+\s-\s\d+\.pdf", f)],
        key=lambda x: int(re.match(r"(\d+)", x).group(1))
    )
    if not pdf_files:
        print("Keine passend benannten PDF-Dateien für die Excel-Erstellung gefunden.")
        return

    for filename in pdf_files:
        pdf_path = os.path.join(dated_folder, filename)
        
        # Initialisiere die Variablen für jeden Patienten
        patientennummer = name = geschlecht = geburtsdatum = ""
        
        # --- START: Überarbeiteter Extraktions-Block ---
        
        diagnose = ""
        # Definiere den Ankertext, der für die Positionierung entscheidend ist
        anchor_text = "Ergänzungen und neue Inhalte werden nicht automatisch in die Tumordokumentation zurückgeschrieben"
        
        try:
            # --- Extraktion der Stammdaten (Name, Nummer etc.) ---
            # Dieser Teil bleibt gleich und nutzt die schnelle Text-Extraktion
            with fitz.open(pdf_path) as doc:
                text = ""
                for page in doc:
                    text += page.get_text()
            
            idx_info = text.find("Patienteninformationen")
            if idx_info != -1:
                after_info = text[idx_info + len("Patienteninformationen"):].lstrip()
                match_num = patient_num_re.search(after_info)
                if match_num:
                    patientennummer = match_num.group()
                    lines = after_info[match_num.end():].lstrip().splitlines()
                    if lines:
                        name = lines[0].strip()
                match_geschlecht = geschlecht_geb_re.search(after_info)
                if match_geschlecht:
                    geschlecht = match_geschlecht.group(1)
                    geburtsdatum = match_geschlecht.group(2)

            # --- Extraktion der DIAGNOSE (mit Fall-back-Logik) ---

            # --- VERSUCH 1: Schnelle Text-Extraktion für Diagnose ---
            print(f"\n[{os.path.basename(filename)}] Versuch 1: Schnelle Text-Extraktion der Diagnose...")
            
            idx_anchor = text.find(anchor_text)
            if idx_anchor != -1:
                text_after_anchor = text[idx_anchor + len(anchor_text):]
                idx_diag_heading = text_after_anchor.lower().find("diagnose")
                if idx_diag_heading != -1:
                    text_after_heading = text_after_anchor[idx_diag_heading + len("diagnose"):].lstrip()
                    lines = text_after_heading.splitlines()
                    for line in lines:
                        cleaned_line = line.strip()
                        if cleaned_line:
                            diagnose = cleaned_line
                            break

            # --- VALIDIERUNG durch Zählen der Buchstaben ---
            # Entferne alle Nicht-Buchstaben und prüfe die Länge des Ergebnisses.
            letter_count = len(re.sub(r'[^a-zA-Z]', '', str(diagnose)))

            if letter_count >= 15: # Schwellenwert für eine "sinnvolle" Diagnose
                print(f"[{os.path.basename(filename)}] Erfolg mit schneller Extraktion. Diagnose: '{diagnose}'")
            else:
                print(f"[{os.path.basename(filename)}] Schnelle Extraktion lieferte keine sinnvolle Diagnose ('{diagnose}'). Anzahl Buchstaben: {letter_count}.")
                print(f"[{os.path.basename(filename)}] VERSUCH 2: Fall-back auf OCR...")
                
                # --- VERSUCH 2: Langsame OCR-Extraktion (Fall-back) ---
                ocr_text = ""
                with fitz.open(pdf_path) as doc:
                    for page_num, page in enumerate(doc):
                        pix = page.get_pixmap(dpi=300)
                        pil_image = Image.open(io.BytesIO(pix.tobytes("png")))
                        
                        temp_image_path = os.path.join(dated_folder, f"temp_ocr_for_{os.path.basename(filename)}.png")
                        pil_image.save(temp_image_path)

                        ocr_result_page = UNIVERSAL.run_tesseract_ocr_deutsch(temp_image_path)
                        
                        os.remove(temp_image_path)

                        if ocr_result_page:
                            ocr_text += ocr_result_page + "\n"
                
                # Diagnose aus dem OCR-Text extrahieren
                diagnose = "" # Reset
                idx_anchor_ocr = ocr_text.find(anchor_text)
                if idx_anchor_ocr != -1:
                    text_after_anchor_ocr = ocr_text[idx_anchor_ocr + len(anchor_text):]
                    idx_diag_heading_ocr = text_after_anchor_ocr.lower().find("diagnose")
                    if idx_diag_heading_ocr != -1:
                        text_after_heading_ocr = text_after_anchor_ocr[idx_diag_heading_ocr + len("diagnose"):].lstrip()
                        lines_ocr = text_after_heading_ocr.splitlines()
                        for line in lines_ocr:
                            cleaned_line_ocr = line.strip().replace('e ', '• ')
                            if cleaned_line_ocr:
                                diagnose = cleaned_line_ocr
                                break
                
                # --- FINALE VALIDIERUNG nach OCR ---
                letter_count_ocr = len(re.sub(r'[^a-zA-Z]', '', str(diagnose)))
                if letter_count_ocr >= 15:
                    print(f"[{os.path.basename(filename)}] Erfolg mit OCR. Diagnose: '{diagnose}'")
                else:
                    print(f"[{os.path.basename(filename)}] FEHLER: Auch OCR lieferte keine sinnvolle Diagnose ('{diagnose}'). Anzahl Buchstaben: {letter_count_ocr}.")
                    diagnose = "MANUELLE PRÜFUNG ERFORDERLICH"

        except Exception as e:
            print(f"[{os.path.basename(filename)}] Ein unerwarteter Fehler ist bei der Extraktion aufgetreten: {e}")
            diagnose = "EXTRAKTIONSFEHLER"
        
        # --- ENDE: Ende des überarbeiteten Extraktions-Blocks ---

        patient_data.append([patientennummer, name, geschlecht, geburtsdatum, diagnose])

    # --- Excel-Template kopieren ---
    
    # 1. Definiere die beiden möglichen Pfade
    net_template_path = r"K:\RAO_Projekte\App\tumorboards\__SQLite_database\template.xlsx"
    local_fallback_template_path = os.path.join(user_home, "tumorboards", "__SQLite_database", "template.xlsx")
    
    template_path = None # Initialisiere den Pfad als None

    # 2. Prüfe zuerst den Netzwerkpfad
    print(f"Suche Template auf dem Netzwerkpfad: {net_template_path}")
    if os.path.exists(net_template_path):
        template_path = net_template_path
        print("Netzwerk-Template gefunden. Verwende dieses.")
    else:
        # 3. Wenn der Netzwerkpfad nicht existiert, prüfe den lokalen Fallback-Pfad
        print(f"Netzwerk-Template nicht gefunden. Prüfe lokalen Fallback-Pfad: {local_fallback_template_path}")
        if os.path.exists(local_fallback_template_path):
            template_path = local_fallback_template_path
            print("Lokales Fallback-Template gefunden. Verwende dieses.")
        else:
            # 4. Wenn beide Pfade fehlschlagen, gib eine Fehlermeldung aus und beende
            print("\nFEHLER: Excel-Template wurde WEDER auf dem Netzwerkpfad NOCH im lokalen Fallback-Verzeichnis gefunden.")
            print(f"Geprüfter Netzwerkpfad: {net_template_path}")
            print(f"Geprüfter lokaler Pfad:  {local_fallback_template_path}")
            print("Skript kann nicht fortfahren. Bitte stellen Sie sicher, dass die template.xlsx an einem der beiden Orte verfügbar ist.")
            return # Beendet die Funktion create_excel_file, da das Template fehlt
    


    excel_dst = os.path.join(dated_folder, "template.xlsx")
    shutil.copy(template_path, excel_dst)

    # --- Excel ausfüllen ---
    wb = load_workbook(excel_dst)
    ws = wb.active
    for idx, (patnum, name, geschlecht, geburtsdatum, diagnose) in enumerate(patient_data, start=2):
        ws[f"A{idx}"] = patnum
        ws[f"B{idx}"] = name
        ws[f"C{idx}"] = geschlecht
        ws[f"D{idx}"] = geburtsdatum
        ws[f"E{idx}"] = diagnose
        
    # --- Sammeln und Anzeigen der Patienten, die manuell geprüft werden müssen ---
    manual_check_needed = []
    for data_row in patient_data:
        pat_num = data_row[0]
        pat_name = data_row[1]
        diag_text = data_row[4]
        if "MANUELLE PRÜFUNG ERFORDERLICH" in diag_text or "EXTRAKTIONSFEHLER" in diag_text:
            manual_check_needed.append(f"{pat_name} ({pat_num})")

    if manual_check_needed:
        print("\n" + "="*60)
        print("ACHTUNG: Bei folgenden Patienten konnte die Diagnose nicht automatisch extrahiert werden:")
        for patient_info in manual_check_needed:
            print(f"  -> {patient_info}")
        print("Bitte überprüfen und korrigieren Sie diese Einträge manuell in der Excel-Datei.")
        print("="*60 + "\n")


    # --- Zweite Umbenennung der PDFs basierend auf den extrahierten Namen ---
    print("\nBenenne PDFs final um (Nachname - Patientennummer.pdf)...")
    for i, filename in enumerate(pdf_files):
        if i < len(patient_data):
            patnum, name, _, _, _ = patient_data[i]
            
            nachname = "Unbekannt"
            if name and name.strip():
                nachname = name.strip().split(',')[0].strip().split()[0]

            if patnum and nachname != "Unbekannt":
                old_path = os.path.join(dated_folder, filename)
                new_filename = f"{nachname} - {patnum}.pdf"
                new_path = os.path.join(dated_folder, new_filename)
                
                if os.path.exists(old_path):
                    try:
                        os.rename(old_path, new_path)
                        print(f"'{filename}' umbenannt in '{new_filename}'")
                    except Exception as e:
                        print(f"Fehler beim Umbenennen von '{filename}' zu '{new_filename}': {e}")
                else:
                    print(f"Warnung: Quelldatei für Umbenennung nicht gefunden: {old_path}")
            else:
                print(f"Warnung: Kein Name oder Patientennummer für '{filename}', Umbenennung wird übersprungen.")

    # --- Excel-Datei speichern und umbenennen ---
    date_str = datetime.now().strftime("%d.%m.%Y")
    excel_final = os.path.join(dated_folder, f"{date_str}.xlsx")
    wb.save(excel_final)
    print(f"\nExcel-Datei gespeichert als {excel_final}")

    # --- Template-Datei im Zielordner löschen ---
    try:
        os.remove(excel_dst)
        print(f"Template-Datei {excel_dst} wurde gelöscht.")
    except Exception as e:
        print(f"Fehler beim Löschen der Template-Datei: {e}")

    return excel_final # für die ICD-Anreicherung



def icd(excel_path):
    # Prüft, ob die Funktion für den aktuellen Benutzer ausgeführt werden soll.
    if not user_home.endswith("votma"):
        return
    
    print("lese k1 und k2...")
    try:
        with open(os.path.join(user_home, 'kp', 'k1.txt'), 'r') as f:
            k1 = f.read().strip()
        with open(os.path.join(user_home, 'kp', 'k2.txt'), 'r') as f:
            k2 = f.read().strip()
        with open(os.path.join(user_home, 'kp', 'p_dg.txt'), 'r') as f:
            p_dg = f.read().strip()
        from google import genai
    except Exception as e:
        print(f"Fehler beim Lesen der Dateien aus dem kp Ordner: {e}")
        return False
    k = str(k1) + str(k2)


    # Lade die Excel-Datei
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
    except FileNotFoundError:
        print(f"Fehler: Excel-Datei nicht gefunden unter {excel_path}")
        return
    except Exception as e:
        print(f"Fehler beim Laden der Excel-Datei: {e}")
        return

    # 1. Sammle alle Diagnosen und ihre Zeilennummern
    diagnoses_list = []
    row_indices = []
    for row_index in range(2, ws.max_row + 1):
        diagnosis_cell = ws[f'E{row_index}']
        if diagnosis_cell.value and str(diagnosis_cell.value).strip():
            diagnoses_list.append(str(diagnosis_cell.value).strip())
            row_indices.append(row_index)
    
    if not diagnoses_list:
        print("Keine Diagnosen in der Excel-Datei gefunden. Überspringe req...")
        return

    print(f"{len(diagnoses_list)} Diagnosen gefunden. Bereite einzelne req vor...")

    # 2. Erstelle einen einzigen, gebündelten Prompt
    prompt = (
        f"Gegeben ist eine Python-Liste von medizinischen Diagnosen: {diagnoses_list}. "
        "Für JEDE Diagnose in dieser Liste, gib den wahrscheinlichsten deutschen ICD-10-GM-Code und die dazugehörige offizielle deutsche Beschreibung zurück. "
        "Für Tumore und tumorverdächtige Raumforderungen dürfen ausschliesslich C- und D-Diagnosen kodiert werden."
        "Bei einem Verdacht auf ein Malignom kann bereits die entsprechende C-Diagnose verwendet werden (z.B. C34.1 hochsuspekte Raumforderung im Oberlappen), bei benignen Tumoren entsprechend die D-Diagnose. Die ausgewählten Codes sollen entsprechend der beschreibenden Lokalisation möglichst präzise sein."
        "Diagnosen, die keiner Tumorerkrankung entsprechen, dürfen durch ihren jeweiligen ICD-10-GM-Code kodiert werden."
        "Formatiere die gesamte Antwort als ein einziges JSON-Objekt, das eine Liste von Objekten ist. "
        "Die zurückgegebene Liste MUSS exakt die gleiche Anzahl an Elementen haben wie die Eingabeliste und die Reihenfolge beibehalten. "
        "Jedes Objekt in der Liste muss die Schlüssel 'icd_code' und 'icd_beschreibung' enthalten. "
        "Gib NUR das JSON-Objekt zurück, ohne zusätzlichen Text, Erklärungen oder Markdown-Formatierung wie ```json ... ```."
        "Beispiel für eine erwartete Antwort bei einer Liste mit zwei Eingabediagnosen: "
        "[{\"icd_code\": \"C50.9\", \"icd_beschreibung\": \"Bösartige Neubildung der Brustdrüse, nicht näher bezeichnet\"}, "
        "{\"icd_code\": \"C18.7\", \"icd_beschreibung\": \"Bösartige Neubildung des Colon sigmoideum\"}]"
    )

    try:
        # 3. Sende die gebündelte Anfrage
        
        c = genai.Client(api_key=f"{k}")
        print("icd req...")
        response = c.models.generate_content(
            model="gemini-2.5-flash", contents=f"{prompt}"
        )
        response_text = response.text.strip().replace("```json", "").replace("```", "").strip()
        results_data = json.loads(response_text)

        # 4. Überprüfe die Antwort und schreibe die Ergebnisse zurück
        if isinstance(results_data, list) and len(results_data) == len(diagnoses_list):
            print("icd req done...")
            for i, result in enumerate(results_data):
                row = row_indices[i]
                icd_code = result.get("icd_code", "N/A")
                icd_beschreibung = result.get("icd_beschreibung", "Keine Beschreibung gefunden")
                
                ws[f'F{row}'] = icd_code
                ws[f'G{row}'] = icd_beschreibung
            
            # 5. Speichere die aktualisierte Excel-Datei
            wb.save(excel_path)
            print(f"\nICD-10-Anreicherung abgeschlossen. Datei '{excel_path}' wurde aktualisiert.")

        else:
            print("\nFEHLER: req hatte ein unerwartetes Format oder eine falsche Anzahl von Ergebnissen.")
            print(f"Erwartet: Liste mit {len(diagnoses_list)} Elementen. Erhalten: {type(results_data)} mit {len(results_data) if isinstance(results_data, list) else 'N/A'} Elementen.")

    except json.JSONDecodeError:
        print(f"\nFEHLER: Kein JSON zurück. Antwortbeginn: {response.text[:200]}...")
    except Exception as e:
        print(f"\nFEHLER bei der Anfrage: {e}")


def main():
    # Check if the script is running in the correct environment
    if not os.path.exists(user_home):
        print("Fehler: Benutzerverzeichnis nicht gefunden.")
        sys.exit(1)

    # Create the screenshots directory if it doesn't exist
    if not os.path.exists(local_screenshots_dir):
        os.makedirs(local_screenshots_dir)

    # Get user input for tumorboard selection
    userinput()

    #Check, dass KISIM im Vordergrund ist
    if not UNIVERSAL.KISIM_im_vordergrund():
        print("Fehler: KISIM ist nicht im Vordergrund.")
        sys.exit(1)

    
    open_tumorboard()
    tumorboard_nach_nachname_sortieren()
    als_pdf_speichern()
    # Nach dem Verschieben der PDFs aufrufen:
    pdfs_umbenennen(dated_folder)
    excel_path = create_excel_file()
    
    # Wenn die Excel-Datei erfolgreich erstellt wurde, starte die ICD-10-Anreicherung
    if excel_path:
        print(f"excel_path: {excel_path}; start icd...")
        icd(excel_path)

    print("\nProzess abgeschlossen.")


if __name__ == "__main__":
    main()