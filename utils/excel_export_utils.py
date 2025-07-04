import pandas as pd
import logging
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import shutil
from datetime import datetime
from .database_utils import TumorboardDatabase




def export_tumorboard_to_collection(tumorboard_name, date_str, tumorboard_base_path=None, skip_database_sync=False):
    """Export a specific tumorboard session to the collection Excel file"""
    try:
        # Use provided base path or fallback to user home
        if tumorboard_base_path is None:
            tumorboard_base_path = Path.home() / "tumorboards"
            
        # Paths
        tumorboard_dir = tumorboard_base_path / tumorboard_name
        daily_excel_path = tumorboard_dir / date_str / f"{date_str}.xlsx"
        
        # Find collection file flexibly - look for any file starting with "alle_tumorboards_"
        collection_file_path = None
        for file in tumorboard_dir.glob("alle_tumorboards_*.xlsx"):
            collection_file_path = file
            break
        
        # Check if daily Excel file exists
        if not daily_excel_path.exists():
            logging.error(f"Daily Excel file not found: {daily_excel_path}")
            return False
        
        # Check if collection file exists
        if not collection_file_path or not collection_file_path.exists():
            logging.error(f"Collection Excel file not found in {tumorboard_dir}. Looking for files starting with 'alle_tumorboards_'")
            return False
        
        # Read the daily Excel data
        daily_df = pd.read_excel(daily_excel_path, engine='openpyxl')
        
        # Load or create collection workbook
        try:
            collection_wb = load_workbook(collection_file_path)
        except Exception as e:
            logging.warning(f"Could not load existing collection file, creating new one: {e}")
            collection_wb = Workbook()
            # Remove default sheet if it exists
            if "Sheet" in collection_wb.sheetnames:
                collection_wb.remove(collection_wb["Sheet"])
        
        # Ensure the overview sheet exists and is properly named
        if "Tabelle1" in collection_wb.sheetnames:
            overview_sheet = collection_wb["Tabelle1"]
            overview_sheet.title = "Übersicht"
        elif "Übersicht" not in collection_wb.sheetnames:
            # Create overview sheet if it doesn't exist
            overview_sheet = collection_wb.create_sheet("Übersicht", 0)
            overview_sheet['A1'] = f"Sammel-Excel für Tumorboard: {tumorboard_name}"
            overview_sheet['A2'] = "Erstellt automatisch durch USZ-RAO-App"
            overview_sheet['A3'] = "Jeder Tab repräsentiert ein Tumorboard-Datum"
        
        # Check if sheet with this date already exists
        sheet_name = date_str.replace(".", "_")  # Excel sheet names can't contain dots
        
        if sheet_name in collection_wb.sheetnames:
            # Remove existing sheet to overwrite
            collection_wb.remove(collection_wb[sheet_name])
            logging.info(f"Removed existing sheet '{sheet_name}' for overwrite")
        
        # Create new sheet
        new_sheet = collection_wb.create_sheet(title=sheet_name)
        
        # Write data to new sheet
        for r in dataframe_to_rows(daily_df, index=False, header=True):
            new_sheet.append(r)
        
        # Auto-adjust column widths
        for column in new_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            new_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save collection workbook
        collection_wb.save(collection_file_path)
        
        logging.info(f"Successfully exported {tumorboard_name} {date_str} to collection Excel")
        
        # After successful collection Excel export, sync to central database
        # Skip database sync if using fallback path (not K: drive)
        if not skip_database_sync:
            try:
                sync_collection_to_database(tumorboard_name, collection_file_path, tumorboard_base_path)
            except Exception as db_error:
                logging.warning(f"Collection Excel export succeeded, but database sync failed: {db_error}")
                # Don't fail the entire operation if database sync fails
        else:
            logging.info(f"Skipping database sync for {tumorboard_name} {date_str} - using fallback path")
        
        return True
        
    except Exception as e:
        logging.error(f"Error exporting tumorboard to collection: {e}")
        return False


def sync_collection_to_database(tumorboard_name, collection_excel_path, tumorboard_base_path=None):
    """Sync a specific collection Excel file to the central database"""
    try:
        # Determine correct database path based on tumorboard base path
        if tumorboard_base_path is not None:
            db_path = tumorboard_base_path / "__SQLite_database" / "master_tumorboard.db"
        else:
            db_path = None  # Use default user home path
        
        db = TumorboardDatabase(db_path=db_path)
        success = db.import_collection_excel(tumorboard_name, collection_excel_path, create_backup=True)
        
        if success:
            logging.info(f"Successfully synced {tumorboard_name} collection to central database")
        else:
            logging.warning(f"Failed to sync {tumorboard_name} collection to central database")
        
        return success
        
    except Exception as e:
        logging.error(f"Error syncing collection to database for {tumorboard_name}: {e}")
        return False


def sync_all_collections_to_database():
    """Sync all collection Excel files to the central database"""
    try:
        from .database_utils import sync_all_collection_files
        return sync_all_collection_files()
        
    except Exception as e:
        logging.error(f"Error syncing all collections to database: {e}")
        return False


def create_category_backup(category_file_path, category):
    """Create a timestamped backup of a category Excel file before modifying it"""
    try:
        if not category_file_path.exists():
            # No file to backup
            return True
            
        # Create backup directory structure
        backoffice_dir = category_file_path.parent
        backup_dir = backoffice_dir / "backup"
        category_backup_dir = backup_dir / category.replace(" ", "_")
        
        # Ensure backup directories exist
        category_backup_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate timestamp for backup filename
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        backup_filename = f"{timestamp}_{category.replace(' ', '_')}.xlsx"
        backup_file_path = category_backup_dir / backup_filename
        
        # Copy the file to backup location
        shutil.copy2(category_file_path, backup_file_path)
        
        logging.info(f"Created backup: {backup_file_path}")
        return True
        
    except Exception as e:
        logging.error(f"Error creating backup for {category}: {e}")
        return False


def export_patients_by_category(tumorboard_name, date_str, tumorboard_base_path=None):
    """Export patients to category-specific backoffice Excel files based on 'Art des Aufgebots'"""
    try:
        # Use provided base path or fallback to user home
        if tumorboard_base_path is None:
            tumorboard_base_path = Path.home() / "tumorboards"
            
        # Read the tumorboard data
        tumorboard_dir = tumorboard_base_path / tumorboard_name
        daily_excel_path = tumorboard_dir / date_str / f"{date_str}.xlsx"
        
        if not daily_excel_path.exists():
            logging.error(f"Daily Excel file not found: {daily_excel_path}")
            return False
        
        # Read the daily Excel data
        daily_df = pd.read_excel(daily_excel_path, engine='openpyxl')
        
        # Ensure backoffice directory exists - use the same base path as the tumorboard
        backoffice_dir = tumorboard_base_path / "_Backoffice"
        backoffice_dir.mkdir(parents=True, exist_ok=True)
        
        # Category mappings
        categories = {
            "Kat I": "Kat_I.xlsx",
            "Kat II": "Kat_II.xlsx", 
            "Kat III": "Kat_III.xlsx"
        }
        
        # Today's date for column A
        today_date = datetime.now().strftime("%d.%m.%Y")
        
        exported_count = 0
        
        # Process each category
        for category, filename in categories.items():
            # Filter patients by category
            category_patients = []
            
            for _, row in daily_df.iterrows():
                # First check if radiotherapy is indicated
                radiotherapy_value = str(row.get('Radiotherapie indiziert', '')).strip()
                if radiotherapy_value != 'Ja':
                    continue  # Skip patients without radiotherapy indication
                
                aufgebot_value = str(row.get('Art des Aufgebots', '')).strip()
                
                # Use precise category matching to avoid false positives
                category_match = False
                
                if category == "Kat I" and aufgebot_value.startswith("Kat I:"):
                    category_match = True
                elif category == "Kat II" and aufgebot_value.startswith("Kat II:"):
                    category_match = True
                elif category == "Kat III" and aufgebot_value.startswith("Kat III:"):
                    category_match = True
                
                # Check if this patient belongs to current category
                if category_match:
                    # Prepare patient data for export (correct column mapping)
                    patient_row = [
                        today_date,  # Column A: Heutiges Datum
                        tumorboard_name,  # Column B: Tumorboard-Klasse
                        str(row.get('Name', '')),  # Column C: Name
                        str(row.get('Geburtsdatum', '')),  # Column D: Geburtsdatum
                        str(row.get('Patientennummer', '')),  # Column E: Patientennummer
                        str(row.get('Diagnose', '')),  # Column F: Diagnose
                        str(row.get('ICD-10', row.get('ICD-Code', ''))),  # Column G: ICD-Code
                        str(row.get('Radiotherapie indiziert', '')),  # Column H: Radiotherapie indiziert
                        str(row.get('Art des Aufgebots', '')),  # Column I: Art des Aufgebots
                        str(row.get('Teams Priorisierung', '')),  # Column J: Teams Priorisierung
                        str(row.get('Vormerken für Studie', '')),  # Column K: Vormerken für Studie
                        str(row.get('Bemerkung/Procedere', ''))  # Column L: Bemerkung/Procedere
                    ]
                    category_patients.append(patient_row)
            
            # Only proceed if there are patients for this category
            if not category_patients:
                continue
            
            # Process the category file
            category_file_path = backoffice_dir / filename
            
            # Create backup before modifying the file
            backup_success = create_category_backup(category_file_path, category)
            if not backup_success:
                logging.warning(f"Backup creation failed for {category}, but continuing with export")
                # Continue with export even if backup fails
            
            # Load existing workbook if it exists, otherwise create new one
            if category_file_path.exists():
                try:
                    wb = load_workbook(category_file_path)
                    ws = wb.active
                    logging.info(f"Loaded existing {filename} with formatting preserved")
                    
                    # Completely remove all Excel Table definitions to avoid corruption
                    if hasattr(ws, 'tables') and ws.tables:
                        tables_to_remove = list(ws.tables.keys())
                        for table_name in tables_to_remove:
                            del ws.tables[table_name]
                            logging.info(f"Removed table definition '{table_name}' to prevent corruption")
                        # Clear the tables collection completely
                        ws.tables.clear()
                    
                except Exception as e:
                    logging.warning(f"Could not load existing {filename}, creating new one: {e}")
                    wb = Workbook()
                    ws = wb.active
                    # Set proper headers (columns A-N) only if creating new file
                    headers = [
                        'Datum',                    # A
                        'Tumorboard',              # B  
                        '',                        # C (bleibt frei)
                        'Name',                    # D
                        'Geburtsdatum',            # E
                        'Patientennummer',         # F
                        'Diagnose',                # G
                        'ICD-Code',                # H
                        'Radiotherapie indiziert', # I
                        'Art des Aufgebots',       # J
                        'Teams Priorisierung',     # K
                        'Vormerken für Studie',    # L
                        'Bemerkung/Procedere',     # M
                        'Status'                   # N (zusätzliche Spalte)
                    ]
                    ws.append(headers)
            else:
                # Create new workbook with headers
                wb = Workbook()
                ws = wb.active
                headers = [
                    'Datum',                   # A
                    'Tumorboard',              # B  
                    '',                        # C (bleibt frei)
                    'Name',                    # D
                    'Geburtsdatum',            # E
                    'Patientennummer',         # F
                    'Diagnose',                # G
                    'ICD-Code',                # H
                    'Radiotherapie indiziert', # I
                    'Art des Aufgebots',       # J
                    'Teams Priorisierung',     # K
                    'Vormerken für Studie',    # L
                    'Bemerkung/Procedere',     # M
                    'Status'                   # N (zusätzliche Spalte)
                ]
                ws.append(headers)
            
            # Insert new rows at the top (row 2) to keep newest entries visible
            num_new_patients = len(category_patients)
            
            # Insert new rows at position 2 (pushes existing data down)
            ws.insert_rows(2, num_new_patients)
            logging.info(f"Inserted {num_new_patients} new rows at position 2 for {category}")
            
            # Add new category patients to the inserted rows (starting from row 2)
            for idx, patient_row in enumerate(category_patients):
                target_row = 2 + idx
                
                # Ensure we have enough columns (add empty Status column for column N)
                patient_row_extended = patient_row + ['']  # Add empty Status column
                
                # Write to specific row and columns A through N (avoid column O)
                for col_idx, value in enumerate(patient_row_extended, start=1):
                    if col_idx <= 14:  # Only write to columns A-N, preserve column O
                        ws.cell(row=target_row, column=col_idx, value=value)
            
            # Create dropdown validation for new rows in column O
            try:
                from openpyxl.worksheet.datavalidation import DataValidation
                
                # Create a new dropdown validation for Ja/Nein
                dv = DataValidation(type="list", formula1='"Ja,Nein"', allow_blank=True)
                dv.error = 'Your entry is not in the list'
                dv.errorTitle = 'Invalid Entry'
                dv.prompt = 'Please select from the list'
                dv.promptTitle = 'List Selection'
                
                # Add the validation to the worksheet
                ws.add_data_validation(dv)
                
                # Apply validation to new rows in column O
                for idx in range(num_new_patients):
                    target_row = 2 + idx
                    target_cell = ws.cell(row=target_row, column=15)  # Column O
                    target_cell.value = "Nein"  # Set default value
                    dv.add(target_cell)
                
                logging.info(f"Created dropdown validation for {num_new_patients} new rows in column O with default 'Nein'")
                
            except Exception as validation_error:
                logging.warning(f"Could not create dropdown validation: {validation_error}")
                # Continue without validation - not critical for data integrity
            
            # DO NOT create new table definitions - work with normal cells only
            # This prevents table corruption issues while preserving data and basic formatting
            logging.info(f"Added {len(category_patients)} patients without table definitions to prevent corruption")
            
            # Save the workbook (preserves existing formatting)
            wb.save(category_file_path)
            
            exported_count += len(category_patients)
            logging.info(f"Exported {len(category_patients)} {category} patients to {filename}")
        
        logging.info(f"Successfully exported {exported_count} patients to category backoffice files")
        return True
        
    except Exception as e:
        logging.error(f"Error exporting patients by category: {e}")
        return False

 