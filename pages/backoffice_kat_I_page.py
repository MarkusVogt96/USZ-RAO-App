from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
                             QFrame, QMessageBox, QTableWidget, QTableWidgetItem, 
                             QHeaderView, QProgressBar, QScrollArea, QComboBox,
                             QDateEdit, QDialog, QDialogButtonBox)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QDate, QTimer
from PyQt6.QtGui import QFont
import os
import logging
import pandas as pd
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook
import getpass

# Import centralized path management
from utils.path_management import BackofficePathManager

class BackofficeKatIPage(QWidget):
    def __init__(self, main_window):
        super().__init__()
        logging.info("Initializing BackofficeKatIPage...")
        self.main_window = main_window
        self.category_name = "Kategorie I"
        self.excel_filename = "Kat_I.xlsx"
        self.patients_data = []
        
        # Column visibility configuration
        # User can customize which columns to show/hide
        self.column_visibility = {
            0: True,   # Datum
            1: True,   # Tumorboard  
            2: True,   # Name
            3: True,   # Geburtsdatum
            4: True,   # Patientennummer
            5: True,   # Diagnose
            6: True,   # ICD-Code
            7: False,  # Radiotherapie indiziert (hidden per user request)
            8: False,  # Art des Aufgebots (hidden per user request)
            9: True,   # Teams Priorisierung
            10: False, # Vormerken für Studie (hidden per user request)
            11: False, # Bemerkung/Procedere (hidden per user request)
            12: True,  # Timestamp
            13: True,  # Bearbeitet (Aktionen)
        }
        
        self.setup_ui()
        self.load_data()
        logging.info("BackofficeKatIPage initialization complete.")

    def setup_ui(self):
        """Setup the category page user interface"""
        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(30, 20, 30, 30)
        main_layout.setSpacing(20)
        self.setLayout(main_layout)

        # Header with title
        header_layout = QHBoxLayout()
        header_layout.setSpacing(15)

        # Title
        title_label = QLabel(f"📋 {self.category_name} - Aufgebot in 1-3 Tagen")
        title_label.setFont(QFont("Helvetica", 24, QFont.Weight.Bold))
        title_label.setStyleSheet("color: white;")
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Refresh button
        self.refresh_button = QPushButton("🔄 Aktualisieren")
        self.refresh_button.setFont(QFont("Helvetica", 12))
        self.refresh_button.setFixedSize(170, 40)  # Made 30px wider
        self.refresh_button.setStyleSheet("""
            QPushButton {
                background-color: #114473;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1a5a9e;
            }
            QPushButton:pressed {
                background-color: #0d3355;
            }
        """)
        self.refresh_button.clicked.connect(self.refresh_data)
        header_layout.addWidget(self.refresh_button)

        main_layout.addLayout(header_layout)

        # Content
        self.create_content(main_layout)

    def create_content(self, parent_layout):
        """Create the main content"""
        # Scrollable area for content
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                background-color: #232F3B;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #425061;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #5A6B7D;
            }
        """)
        
        # Content widget inside scroll area
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setSpacing(25)
        content_layout.setContentsMargins(0, 0, 0, 0)

        # Filter section
        self.create_filter_section(content_layout)

        # Table section
        self.create_table_section(content_layout)

        # Set scroll area content
        scroll_area.setWidget(content_widget)
        parent_layout.addWidget(scroll_area)

    def create_filter_section(self, parent_layout):
        """Create filter controls"""
        filter_frame = QFrame()
        filter_frame.setStyleSheet("""
            QFrame {
                background-color: #1a2633;
                border: 1px solid #425061;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        
        filter_layout = QHBoxLayout(filter_frame)
        filter_layout.setSpacing(20)

        # Tumorboard filter
        tumorboard_label = QLabel("Tumorboard:")
        tumorboard_label.setFont(QFont("Helvetica", 12))
        tumorboard_label.setStyleSheet("color: white;")
        filter_layout.addWidget(tumorboard_label)

        self.tumorboard_filter = QComboBox()
        self.tumorboard_filter.setStyleSheet("""
            QComboBox {
                background-color: #232F3B;
                border: none;
                outline: none;
                border-radius: 4px;
                padding: 5px;
                color: white;
                min-width: 150px;
            }
            QComboBox::drop-down {
                border: none;
                outline: none;
            }
            QComboBox::down-arrow {
                width: 12px;
                height: 12px;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #425061;
                outline: none;
            }
        """)
        self.tumorboard_filter.currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.tumorboard_filter)

        filter_layout.addSpacing(30)

        # Date range filter
        date_label = QLabel("Datum ab:")
        date_label.setFont(QFont("Helvetica", 12))
        date_label.setStyleSheet("color: white;")
        filter_layout.addWidget(date_label)

        self.date_filter = QDateEdit()
        self.date_filter.setDate(QDate.currentDate().addDays(-30))  # Default: last 30 days
        self.date_filter.setStyleSheet("""
            QDateEdit {
                background-color: #232F3B;
                border: none;
                outline: none;
                border-radius: 4px;
                padding: 5px;
                color: white;
                min-width: 120px;
            }
            QDateEdit::drop-down {
                border: none;
                outline: none;
            }
        """)
        self.date_filter.dateChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.date_filter)

        filter_layout.addStretch()
        parent_layout.addWidget(filter_frame)

    def create_table_section(self, parent_layout):
        """Create the main table section"""
        # Table frame (same styling as Leistungsabrechnungen)
        table_frame = QFrame()
        table_frame.setStyleSheet("""
            QFrame {
                background-color: transparent;
                border: none;
                padding: 0px;
            }
        """)
        
        table_layout = QVBoxLayout(table_frame)
        
        # Create table
        self.patients_table = QTableWidget()
        self.patients_table.setColumnCount(14)  # A-M + Action column (14 total)
        self.patients_table.setHorizontalHeaderLabels([
            "Datum", "Tumorboard", "Name", "Geburtsdatum", "Patienten-Nr.", 
            "Diagnose", "ICD-Code", "Radiotherapie indiziert", "Art des Aufgebots",
            "Teams Priorisierung", "Vormerken für Studie", "Bemerkung/Procedere", "Timestamp", "Bearbeitet"
        ])
        
        # Apply exact same styling as Leistungsabrechnungen table
        self.patients_table.setStyleSheet("""
            QTableWidget {
                background-color: #232F3B;
                alternate-background-color: #2A3642;
                selection-background-color: #3292ea;
                gridline-color: #425061;
                color: white;
                border: 1px solid #425061;
            }
            QTableWidget::item {
                padding: 3px;
                border-bottom: 1px solid #425061;
            }
            QTableWidget::item:selected {
                background-color: #3292ea;
            }
            QHeaderView::section {
                background-color: #1a2633;
                color: white;
                padding: 3px;
                border: 1px solid #425061;
                font-weight: bold;
                font-size: 16px;
            }
            QTableWidget QTableCornerButton::section {
                background-color: #1a2633;
                border: 1px solid #425061;
            }
        """)
        
        # Set table properties (same as Leistungsabrechnungen)
        self.patients_table.setAlternatingRowColors(True)
        self.patients_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.patients_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.patients_table.setSortingEnabled(True)

        # Configure headers - make all columns resizable by dragging
        horizontal_header = self.patients_table.horizontalHeader()
        
        # Set ALL columns to Interactive mode (resizable by dragging)
        for col in range(14):
            horizontal_header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
        
        # Set some reasonable default widths
        horizontal_header.resizeSection(0, 100)   # Datum
        horizontal_header.resizeSection(1, 120)   # Tumorboard
        horizontal_header.resizeSection(2, 157)   # Name
        horizontal_header.resizeSection(3, 120)   # Geburtsdatum
        horizontal_header.resizeSection(4, 110)   # Patientennummer
        horizontal_header.resizeSection(5, 250)   # Diagnose
        horizontal_header.resizeSection(6, 100)   # ICD-Code
        horizontal_header.resizeSection(7, 150)   # Radiotherapie indiziert
        horizontal_header.resizeSection(8, 130)   # Art des Aufgebots
        horizontal_header.resizeSection(9, 280)   # Teams Priorisierung
        horizontal_header.resizeSection(10, 140)  # Vormerken für Studie
        horizontal_header.resizeSection(11, 180)  # Bemerkung/Procedere
        horizontal_header.resizeSection(12, 150)  # Timestamp
        horizontal_header.resizeSection(13, 150)  # Bearbeitet
        
        # Don't stretch last section to allow custom widths
        horizontal_header.setStretchLastSection(False)

        vertical_header = self.patients_table.verticalHeader()
        vertical_header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        vertical_header.setMinimumSectionSize(60)
        
        # Hide the vertical header (row numbers) to avoid offset issues
        vertical_header.setVisible(False)
        
        # Hide columns based on visibility configuration
        self.apply_column_visibility()
        
        # Fix corner button styling after widget is created
        QTimer.singleShot(100, self._fix_corner_button_styling)
        
        table_layout.addWidget(self.patients_table)
        parent_layout.addWidget(table_frame)

    def _fix_corner_button_styling(self):
        """Fix the corner button styling after the widget is fully initialized"""
        try:
            # Find the corner button and apply styling
            corner_button = self.patients_table.findChild(QPushButton)
            if corner_button:
                corner_button.setStyleSheet("""
                    QPushButton {
                        background-color: #1a2633;
                        border: 1px solid #425061;
                    }
                """)
                # Disable the corner button to prevent select all functionality
                corner_button.setEnabled(False)
                logging.info("Corner button styling applied successfully")
        except Exception as e:
            logging.warning(f"Could not apply corner button styling: {e}")

    def apply_column_visibility(self):
        """Apply column visibility settings - hide/show columns based on configuration"""
        for col_index, is_visible in self.column_visibility.items():
            if col_index < self.patients_table.columnCount():
                if is_visible:
                    self.patients_table.showColumn(col_index)
                else:
                    self.patients_table.hideColumn(col_index)
        
        logging.info(f"Applied column visibility settings - Hidden columns: {[col for col, visible in self.column_visibility.items() if not visible]}")

    def load_data(self):
        """Load data from the category Excel file"""
        try:
            # Use centralized path management with network/local priority
            backoffice_dir, using_network = BackofficePathManager.get_backoffice_path(show_warnings=False)
            excel_path = backoffice_dir / self.excel_filename
            
            if not excel_path.exists():
                logging.warning(f"Category Excel file not found: {excel_path}")
                self.patients_data = []
                self.populate_tumorboard_filter()
                self.populate_table()
                return
            
            # Read Excel file
            df = pd.read_excel(excel_path, engine='openpyxl')
            
            if df.empty:
                self.patients_data = []
                self.populate_tumorboard_filter()
                self.populate_table()
                return
            
            # Convert dataframe to list of dictionaries
            self.patients_data = []
            for index, row in df.iterrows():
                patient = {
                    'row_index': index + 2,  # Excel row (1-based + header)
                    'datum': str(row.iloc[0]) if len(row) > 0 else '',
                    'tumorboard': str(row.iloc[1]) if len(row) > 1 else '',
                    'name': str(row.iloc[2]) if len(row) > 2 else '',  # Removed empty column, shifted indices
                    'geburtsdatum': str(row.iloc[3]) if len(row) > 3 else '',
                    'patientennummer': str(row.iloc[4]) if len(row) > 4 else '',
                    'diagnose': str(row.iloc[5]) if len(row) > 5 else '',
                    'icd_code': str(row.iloc[6]) if len(row) > 6 else '',
                    'radiotherapie': str(row.iloc[7]) if len(row) > 7 else '',
                    'art_aufgebot': str(row.iloc[8]) if len(row) > 8 else '',
                    'teams_priorisierung': str(row.iloc[9]) if len(row) > 9 else '',
                    'studie': str(row.iloc[10]) if len(row) > 10 else '',
                    'bemerkung': str(row.iloc[11]) if len(row) > 11 else '',
                    'timestamp': str(row.iloc[12]) if len(row) > 12 and pd.notna(row.iloc[12]) else '',  # Spalte M (index 12)
                    'status': str(row.iloc[13]).strip().lower() if len(row) > 13 else 'nein'  # Spalte N (index 13)
                }
                self.patients_data.append(patient)
            
            logging.info(f"Loaded {len(self.patients_data)} patients from {self.excel_filename}")
            
        except Exception as e:
            logging.error(f"Error loading data from {self.excel_filename}: {e}")
            self.patients_data = []
        
        # Populate UI elements
        self.populate_tumorboard_filter()
        self.populate_table()

    def populate_tumorboard_filter(self):
        """Populate the tumorboard filter dropdown"""
        # Get unique tumorboards from data
        tumorboards = set()
        for patient in self.patients_data:
            if patient['tumorboard']:
                tumorboards.add(patient['tumorboard'])
        
        # Update filter dropdown
        self.tumorboard_filter.clear()
        self.tumorboard_filter.addItem("Alle")
        for tumorboard in sorted(tumorboards):
            self.tumorboard_filter.addItem(tumorboard)

    def populate_table(self):
        """Populate the table with patient data"""
        # Apply filters to get filtered data
        filtered_data = self.get_filtered_data()
        
        self.patients_table.setRowCount(len(filtered_data))
        
        for row, patient in enumerate(filtered_data):
            # Populate data columns - Note: removed 'empty' column as per user changes
            columns = [
                patient['datum'], patient['tumorboard'], patient['name'],
                patient['geburtsdatum'], patient['patientennummer'], patient['diagnose'],
                patient['icd_code'], patient['radiotherapie'], patient['art_aufgebot'],
                patient['teams_priorisierung'], patient['studie'], patient['bemerkung'], patient['timestamp']
            ]
            
            for col, value in enumerate(columns):
                item = QTableWidgetItem(str(value))
                item.setFont(QFont("Helvetica", 11))
                
                # Color coding based on status
                if patient['status'] == 'ja':
                    # Green for completed
                    item.setBackground(Qt.GlobalColor.darkGreen)
                elif patient['status'] == 'nein':
                    # Red for pending
                    item.setBackground(Qt.GlobalColor.darkRed)
                
                self.patients_table.setItem(row, col, item)
            
            # Create action button (column 13, since we added timestamp column)
            action_widget = self.create_action_widget(patient, row)
            self.patients_table.setCellWidget(row, 13, action_widget)

    def create_action_widget(self, patient, row):
        """Create action dropdown widget for table row"""
        # Create dropdown with Ja/Nein options
        status_dropdown = QComboBox()
        status_dropdown.addItem("Nein")
        status_dropdown.addItem("Ja")
        
        # Set current status based on patient data
        current_status = patient['status'].lower().strip()
        if current_status == 'ja':
            status_dropdown.setCurrentText("Ja")
        else:
            status_dropdown.setCurrentText("Nein")
        
        # Style the dropdown - smaller and centered
        status_dropdown.setStyleSheet("""
            QComboBox {
                background-color: #232F3B;
                border: none;
                border-radius: 4px;
                padding: 3px 3px;
                color: white;
                font-weight: bold;
                text-align: center;
                min-width: 60px;
                max-width: 60px;
                height: 30px;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox::down-arrow {
                width: 8px;
                height: 8px;
            }
            QComboBox QAbstractItemView {
                background-color: #232F3B;
                border: 1px solid #425061;
                selection-background-color: #3292ea;
                color: white;
            }
        """)
        
        # Center the dropdown in a widget
        container_widget = QWidget()
        container_layout = QHBoxLayout(container_widget)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.addStretch()
        container_layout.addWidget(status_dropdown)
        container_layout.addStretch()
        
        # Connect change event
        status_dropdown.currentTextChanged.connect(
            lambda new_status, p=patient, r=row: self.on_status_changed(new_status, p, r)
        )
        
        # Apply color based on current status
        self.update_dropdown_color(status_dropdown, current_status)
        
        return container_widget

    def update_dropdown_color(self, dropdown, status):
        """Update dropdown color based on status"""
        if status == 'ja':
            # Green for completed
            dropdown.setStyleSheet(dropdown.styleSheet() + """
                QComboBox {
                    background-color: #2e7d32;
                    color: white;
                }
            """)
        else:
            # Red for pending
                         dropdown.setStyleSheet(dropdown.styleSheet() + """
                 QComboBox {
                     background-color: #d32f2f;
                     color: white;
                 }
             """)

    def on_status_changed(self, new_status, patient, row):
        """Handle status change from dropdown"""
        try:
            old_status = patient['status'].lower()
            new_status_lower = new_status.lower()
            
            # Check if user is trying to revert from Ja to Nein
            if old_status == 'ja' and new_status_lower == 'nein':
                # Show warning dialog with timestamp information
                if not self.confirm_status_revert(patient):
                    # User cancelled - revert dropdown to original value
                    container_widget = self.patients_table.cellWidget(row, 13)
                    if container_widget:
                        dropdown = container_widget.findChild(QComboBox)
                        if dropdown:
                            dropdown.setCurrentText(patient['status'].title())
                    return
                
                # Log the reversion
                self.log_status_reversion(patient)
            
            # Update Excel file using centralized path management
            backoffice_dir, using_network = BackofficePathManager.get_backoffice_path(show_warnings=False)
            excel_path = backoffice_dir / self.excel_filename
            
            if not excel_path.exists():
                QMessageBox.warning(self, "Fehler", f"Excel-Datei nicht gefunden: {excel_path}")
                return
            
            # Load workbook
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Update status in column N (column 14) - the correct status column
            excel_row = patient['row_index']
            ws.cell(row=excel_row, column=14, value=new_status)
            
            # Handle timestamp for Nein -> Ja transition
            if old_status == 'nein' and new_status_lower == 'ja':
                # Create timestamp with date/time and username
                timestamp = self.create_timestamp()
                ws.cell(row=excel_row, column=13, value=timestamp)  # Column M (index 13 in Excel)
                patient['timestamp'] = timestamp
                
                # Update timestamp display in table
                timestamp_item = QTableWidgetItem(timestamp)
                timestamp_item.setFont(QFont("Helvetica", 11))
                if new_status_lower == 'ja':
                    timestamp_item.setBackground(Qt.GlobalColor.darkGreen)
                else:
                    timestamp_item.setBackground(Qt.GlobalColor.darkRed)
                self.patients_table.setItem(row, 12, timestamp_item)
            
            # Clear timestamp when reverting from Ja to Nein
            elif old_status == 'ja' and new_status_lower == 'nein':
                ws.cell(row=excel_row, column=13, value='')  # Clear timestamp
                patient['timestamp'] = ''
                
                # Clear timestamp display in table
                timestamp_item = QTableWidgetItem('')
                timestamp_item.setFont(QFont("Helvetica", 11))
                timestamp_item.setBackground(Qt.GlobalColor.darkRed)
                self.patients_table.setItem(row, 12, timestamp_item)
            
            # Save workbook
            wb.save(excel_path)
            
            logging.info(f"Changed status of patient {patient['name']} to {new_status} in {self.excel_filename}")
            
            # Update patient data in memory
            patient['status'] = new_status.lower()
            
            # Update dropdown color
            container_widget = self.patients_table.cellWidget(row, 13) # Changed from 12 to 13
            if container_widget:
                # Find the dropdown within the container
                dropdown = container_widget.findChild(QComboBox)
                if dropdown:
                    self.update_dropdown_color(dropdown, new_status.lower())
            
            # Update row background color
            for col in range(13):  # Only data columns, not action column
                item = self.patients_table.item(row, col)
                if item:
                    if new_status.lower() == 'ja':
                        item.setBackground(Qt.GlobalColor.darkGreen)
                    else:
                        item.setBackground(Qt.GlobalColor.darkRed)
            
        except Exception as e:
            logging.error(f"Error changing patient status: {e}")
            QMessageBox.critical(
                self, 
                "Fehler", 
                f"Fehler beim Ändern des Status:\n\n{str(e)}"
            )
            # Revert dropdown to original value
            container_widget = self.patients_table.cellWidget(row, 13) # Changed from 12 to 13
            if container_widget:
                dropdown = container_widget.findChild(QComboBox)
                if dropdown:
                    dropdown.setCurrentText(patient['status'].title())
    
    def create_timestamp(self):
        """Create a timestamp string with date/time and username"""
        now = datetime.now()
        date_time = now.strftime("%d.%m.%Y %H:%M:%S")
        try:
            username = getpass.getuser()
        except:
            username = "Unbekannt"
        
        return f"{date_time}\n{username}"
    
    def confirm_status_revert(self, patient):
        """Show confirmation dialog when reverting status from Ja to Nein"""
        # Parse timestamp if available
        timestamp_info = "Unbekannt"
        if patient['timestamp']:
            try:
                lines = patient['timestamp'].split('\n')
                if len(lines) >= 2:
                    date_time = lines[0]
                    username = lines[1]
                    timestamp_info = f"{date_time} durch {username}"
                else:
                    timestamp_info = patient['timestamp']
            except:
                timestamp_info = patient['timestamp']
        
        # Create confirmation dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Bearbeitungsstatus zurücksetzen")
        dialog.setFixedSize(500, 200)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #19232D;
                color: white;
            }
            QLabel {
                color: white;
                font-size: 12px;
            }
            QPushButton {
                background-color: #114473;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 15px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #1a5a9e;
            }
        """)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        
        # Warning message
        warning_label = QLabel(f"Dieser Patient wurde bereits als bearbeitet markiert am:")
        warning_label.setFont(QFont("Helvetica", 12, QFont.Weight.Bold))
        layout.addWidget(warning_label)
        
        # Timestamp info
        timestamp_label = QLabel(timestamp_info)
        timestamp_label.setFont(QFont("Helvetica", 11))
        timestamp_label.setStyleSheet("color: #FFD700; margin-left: 20px;")
        layout.addWidget(timestamp_label)
        
        # Confirmation question
        question_label = QLabel("Sind Sie sicher, dass Sie den Bearbeitungszustand auf Nein zurücksetzen möchten?")
        question_label.setFont(QFont("Helvetica", 11))
        question_label.setWordWrap(True)
        layout.addWidget(question_label)
        
        # Log info
        log_label = QLabel("Dieser Vorgang wird in den programminternen Logs dokumentiert.")
        log_label.setFont(QFont("Helvetica", 10))
        log_label.setStyleSheet("color: #CCCCCC; font-style: italic;")
        layout.addWidget(log_label)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        yes_button = QPushButton("Ja")
        yes_button.clicked.connect(dialog.accept)
        button_layout.addWidget(yes_button)
        
        no_button = QPushButton("Nein")
        no_button.clicked.connect(dialog.reject)
        button_layout.addWidget(no_button)
        
        layout.addLayout(button_layout)
        
        return dialog.exec() == QDialog.DialogCode.Accepted
    
    def log_status_reversion(self, patient):
        """Log status reversion from Ja to Nein"""
        try:
            # Create log entry
            timestamp = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
            username = getpass.getuser()
            category = self.category_name
            
            log_entry = f"[{timestamp}] {category}: Patient {patient['name']} (ID: {patient['patientennummer']}) von Ja→Nein geändert durch {username}\n"
            
            # Determine log file path using centralized path management
            backoffice_dir, using_network = BackofficePathManager.get_backoffice_path(show_warnings=False)
            log_file = backoffice_dir / "erstkons-logs.txt"
            
            # Ensure directory exists
            backoffice_dir.mkdir(parents=True, exist_ok=True)
            
            # Append to log file
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(log_entry)
            
            logging.info(f"Status reversion logged for patient {patient['name']}")
            
        except Exception as e:
            logging.error(f"Error logging status reversion: {e}")
            # Don't show error to user, as this is just logging

    def mark_as_processed(self, row, patient):
        """Mark a patient as processed (change status from Nein to Ja)"""
        try:
            # Update Excel file using centralized path management
            backoffice_dir, using_network = BackofficePathManager.get_backoffice_path(show_warnings=False)
            excel_path = backoffice_dir / self.excel_filename
            
            if not excel_path.exists():
                QMessageBox.warning(self, "Fehler", f"Excel-Datei nicht gefunden: {excel_path}")
                return
            
            # Load workbook
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Update status in column N (column 14) - the correct status column
            excel_row = patient['row_index']
            ws.cell(row=excel_row, column=14, value="Ja")
            
            # Save workbook
            wb.save(excel_path)
            
            logging.info(f"Marked patient {patient['name']} as processed in {self.excel_filename}")
            
            # Refresh data and table
            self.refresh_data()
            
            # Show success message
            QMessageBox.information(
                self, 
                "Erfolgreich bearbeitet", 
                f"Patient {patient['name']} wurde als bearbeitet markiert."
            )
            
        except Exception as e:
            logging.error(f"Error marking patient as processed: {e}")
            QMessageBox.critical(
                self, 
                "Fehler", 
                f"Fehler beim Markieren des Patienten:\n\n{str(e)}"
            )

    def get_filtered_data(self):
        """Get filtered patient data based on current filter settings"""
        filtered_data = self.patients_data.copy()
        
        # Filter by tumorboard
        tumorboard_filter = self.tumorboard_filter.currentText()
        if tumorboard_filter != "Alle":
            filtered_data = [p for p in filtered_data if p['tumorboard'] == tumorboard_filter]
        
        # Filter by date
        qdate = self.date_filter.date()
        filter_date = date(qdate.year(), qdate.month(), qdate.day())
        filtered_data = [p for p in filtered_data if self.is_date_after_filter(p['datum'], filter_date)]
        
        return filtered_data

    def is_date_after_filter(self, date_str, filter_date):
        """Check if patient date is after filter date"""
        try:
            if not date_str or str(date_str).strip() == 'nan':
                return True  # Include patients without date
            
            # Try to parse the date string
            patient_date = datetime.strptime(str(date_str), '%d.%m.%Y').date()
            return patient_date >= filter_date
        except:
            return True  # Include patients with unparseable dates

    def apply_filters(self):
        """Apply current filter settings to the table"""
        self.populate_table()

    def refresh_data(self):
        """Refresh data from Excel file"""
        self.load_data()
        logging.info(f"Data refreshed for {self.category_name}") 