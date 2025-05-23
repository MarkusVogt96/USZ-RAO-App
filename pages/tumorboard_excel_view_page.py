# pages/tumorboard_excel_view_page.py
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QAbstractItemView, QHeaderView, QApplication
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor
import os
import openpyxl # Für direkten Excel-Zugriff

class TumorboardExcelViewPage(QWidget):
    def __init__(self, main_window, date_folder_path, excel_file_name, board_name):
        super().__init__()
        self.main_window = main_window
        self.date_folder_path = date_folder_path # Für find_page_index
        self.excel_file_path = os.path.join(date_folder_path, excel_file_name)
        self.board_name = board_name # Für den Titel
        self.excel_file_date = os.path.splitext(excel_file_name)[0] # Datum aus Dateinamen extrahieren

        self.setup_ui()
        self.load_excel_data()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(10)

        title_text = f"Tumorboard Liste: {self.board_name} - {self.excel_file_date}"
        title_label = QLabel(title_text)
        title_label.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("color: white; margin-bottom: 10px;")
        main_layout.addWidget(title_label)

        self.table_widget = QTableWidget()
        self.table_widget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers) # Read-only
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_widget.setAlternatingRowColors(True) # Für bessere Lesbarkeit
        
        # Styling für die Tabelle
        self.table_widget.setStyleSheet("""
            QTableWidget {
                background-color: #2c3e50; /* Hintergrund der Tabelle */
                color: white; /* Textfarbe in Zellen */
                gridline-color: #425061; /* Farbe der Gitterlinien */
                border: 1px solid #425061;
                border-radius: 5px;
            }
            QTableWidget::item {
                padding: 5px; /* Innenabstand der Zellen */
            }
            QTableWidget::item:selected {
                background-color: #3292ea; /* Hervorhebungsfarbe */
                color: white;
            }
            QHeaderView::section {
                background-color: #1e2a38; /* Hintergrund der Kopfzeile */
                color: white;
                padding: 6px;
                border: 1px solid #425061; /* Rahmen um Kopfzellen */
                font-weight: bold;
                font-size: 11pt;
            }
        """)
        
        # Kopfzeilen-Styling
        self.table_widget.horizontalHeader().setFont(QFont("Arial", 10, QFont.Weight.Bold))
        self.table_widget.verticalHeader().setVisible(False) # Vertikale Kopfzeile (Zeilennummern) ausblenden
        self.table_widget.horizontalHeader().setStretchLastSection(True) # Letzte Spalte füllt den Platz

        main_layout.addWidget(self.table_widget)

    def load_excel_data(self):
        if not os.path.exists(self.excel_file_path):
            error_item = QTableWidgetItem(f"Excel-Datei nicht gefunden: {self.excel_file_path}")
            self.table_widget.setColumnCount(1)
            self.table_widget.setRowCount(1)
            self.table_widget.setItem(0, 0, error_item)
            self.table_widget.setHorizontalHeaderLabels(["Fehler"])
            return

        try:
            workbook = openpyxl.load_workbook(self.excel_file_path, data_only=True) # data_only=True für Werte statt Formeln
            sheet = workbook.active

            self.table_widget.setRowCount(sheet.max_row -1) # -1, da wir die Kopfzeile nicht als Datenzeile zählen
            self.table_widget.setColumnCount(sheet.max_column)

            headers = [cell.value if cell.value is not None else "" for cell in sheet[1]]
            self.table_widget.setHorizontalHeaderLabels(headers)

            for row_idx, row_data in enumerate(sheet.iter_rows(min_row=2, values_only=True)): # Starte ab Zeile 2 für Daten
                for col_idx, cell_data in enumerate(row_data):
                    item = QTableWidgetItem(str(cell_data) if cell_data is not None else "")
                    item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    self.table_widget.setItem(row_idx, col_idx, item)
            
            self.table_widget.resizeColumnsToContents()
            # Ggf. eine Mindestbreite für Spalten setzen oder Stretch-Faktor anpassen
            header = self.table_widget.horizontalHeader()
            for i in range(header.count()):
                 header.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
            if header.count() > 0: # Letzte Sektion strecken, wenn Spalten vorhanden sind
                header.setSectionResizeMode(header.count() -1, QHeaderView.ResizeMode.Stretch)


        except Exception as e:
            error_item = QTableWidgetItem(f"Fehler beim Laden der Excel-Datei: {e}")
            self.table_widget.setColumnCount(1)
            self.table_widget.setRowCount(1)
            self.table_widget.setItem(0, 0, error_item)
            self.table_widget.setHorizontalHeaderLabels(["Fehler"])
            print(f"FEHLER beim Laden von {self.excel_file_path}: {e}")